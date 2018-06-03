import openpyxl
import pandas as pd
import csv
from textblob import TextBlob
import xlsxwriter
import codecs
import xlrd

#wb=openpyxl.load_workbook('axis_bank.xlsx')
#sheet=wb['axis_bank']

workbook = xlrd.open_workbook('axis_bank.xlsx', encoding_override='cp1252')
sheet=workbook.sheet_by_index(0)
    
with open('axis_sentiment.csv','a',newline='') as f:
    writer=csv.DictWriter(f,fieldnames=['Date','Tweet','Sentiment'])
    writer.writeheader()
    for i in range(1,sheet.max_row):
        date=sheet.cell(row=i,column=1).value
        #date=date.split('_at":')[1]
        #date_x=date[1]
        tweet=sheet.cell(row=i,column=2).value
        #tweet=tweet.split('text:')[1]
        analysis=TextBlob(tweet)
        sentiment=analysis.sentiment.polarity
        if sentiment>0:
            polarity=1
        elif sentiment<0:
            polarity=-1
        else:
            polarity=0
        writer.writerow({'Date':date,'Tweet':tweet,'Sentiment':polarity})
    
    
